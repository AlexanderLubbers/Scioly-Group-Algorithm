from openpyxl import load_workbook
import math
event1 = {}
event2 = {}
event3 = {}
event4 = {}
event5 = {}
event6 = {}
event7 = {}
event8 = {}
event9 = {}
event10 = {}
event11 = {}
event12 = {}
event13 = {}
event14 = {}
event15 = {}
event16 = {}
event17 = {}
event18 = {}
event19 = {}
event20 = {}
event21 = {}
event22 = {}
event23 = {}

class Grouping_Algorithm:
    def getMaxGroupSize(self, sheet, eventName):
        counter = 2
        key = "F" + str(counter)
        while True:
            if eventName == sheet[key].value:
                break
            counter+=1
            key = "F" + str(counter)
        key = "J" + str(counter)
        maxGroupSize = sheet[key].value
        maxGroupSize = int(maxGroupSize)
        return maxGroupSize
    def getPreferredTeamates(self, sheet, name):
        counter = 2
        key = "A" + str(counter)
        while True:
            if name == sheet[key].value:
                break
            counter+=1
            key = "A" + str(counter)
        key = "B" + str(counter)
        string = sheet[key].value
        preferredTeamates = string.split(",")
        preferredTeamates = [preferredTeamates.strip() for preferredTeamates in preferredTeamates]
        return preferredTeamates
    def addToTeamDict(self, team, teams, eventName):
        if eventName in teams:
            iterate = True
            counter = 2
            while iterate == True:
                if (eventName + str(counter)) not in teams:
                    iterate = False
                    teams[eventName+str(counter)] = team
                counter+=1
        else:
            teams[eventName] = team
        return teams
    def getWorkbook(self, filename):
        file = "team_data/" + filename + ".xlsx"
        try:
            workbook = load_workbook(filename=file)
        except:
            print("error: could not find file named: " + filename + ".xlsx")
            return False, False, False
        sheet = workbook.active
        return sheet, workbook, file
    def getNumOfTeams(self, sheet):
        nameCounter = 1
        nameKey = "A" + str(nameCounter)
        doneCounting = False
        while doneCounting == False:
            if sheet[nameKey].value == None:
               doneCounting = True
            nameCounter = nameCounter + 1
            nameKey = "A" + str(nameCounter)
        num = nameCounter - 3
        return math.ceil(num / 15)
    def putInTeams(self, sheet, eventList, numTeams):
        teams = {}
        leftovers = []
        for t in eventList:
            participants = []
            teamsCreated = 0
            eventName = ""
            counter = 0
            for i in t.values():
                if counter == 0:
                    eventName = i
                else:
                    participants.append(i)
                counter+=1
            counter = 0
            maxGroupSize = self.getMaxGroupSize(sheet, eventName)
            #handle edge case
            team = []
            if(len(participants) == 0):
                while teamsCreated != numTeams:
                    teams = self.addToTeamDict(team, teams, eventName)
                    teamsCreated+=1
            else:
                team.append(participants[0])
                name = participants[0]
                participants.remove(participants[0])
                preferredTeamates = self.getPreferredTeamates(sheet, name)
                for i in preferredTeamates:
                    if len(participants) == 0 or numTeams == teamsCreated:
                        break
                    for p in participants:
                        if i == p:
                            team.append(p)
                            participants.remove(p)
                            if len(team) == maxGroupSize:
                                teamsCreated+=1
                                teams = self.addToTeamDict(team, teams, eventName)
                                team = []
                    #no one in the preferred teamate list was participating in the event so
                    #the next person in the participant list should be added as long as they exist
                    #and should be added to create a team
                    if len(participants) != 0 and numTeams > teamsCreated:
                        team.append(participants[0])
                        participants.remove(participants[0])
                        if len(team) == maxGroupSize:
                            teamsCreated+=1
                            teams = self.addToTeamDict(team, teams, eventName)
                            team = []
                if numTeams != teamsCreated:
                    while teamsCreated != numTeams:
                        teams = self.addToTeamDict(team, teams, eventName)
                        team = []
                        teamsCreated+=1
                if len(participants) > 0:
                    for people in participants:
                        leftovers.append(people)
        return teams, leftovers                         
    def putInEvent(self, people, eventLookup):
        for key in people:
            events = people[key]
            eventList = events.split(',')
            eventList = [eventList.strip() for eventList in eventList]
            for t in eventList:
                for i in eventLookup:
                    if t == i.get("Event name"):
                        i["person" + str(len(i))] = key
    def sortSeniors(self, sheet):
        dictionary = {}
        counter = 2
        key = "D" + str(counter)
        doneSorting = False
        while doneSorting == False:
            if sheet[key].value == "Sr":
                nameKey = "A" + str(counter)
                eventKey = "C" + str(counter)
                name = sheet[nameKey].value
                events = sheet[eventKey].value
                dictionary[name] = events
            if sheet[key].value == None:
                doneSorting = True
            counter = counter + 1
            key = "D" + str(counter)
        return dictionary
    def sortJuniors(self, sheet):
        dictionary = {}
        counter = 2
        key = "D" + str(counter)
        doneSorting = False
        while doneSorting == False:
            if sheet[key].value == "J":
                nameKey = "A" + str(counter)
                eventKey = "C" + str(counter)
                name = sheet[nameKey].value
                events = sheet[eventKey].value
                dictionary[name] = events
            if sheet[key].value == None:
                doneSorting = True
            counter = counter + 1
            key = "D" + str(counter)
        return dictionary
    def sortSophomores(self, sheet):
        dictionary = {}
        counter = 2
        key = "D" + str(counter)
        doneSorting = False
        while doneSorting == False:
            if sheet[key].value == "S":
                nameKey = "A" + str(counter)
                eventKey = "C" + str(counter)
                name = sheet[nameKey].value
                events = sheet[eventKey].value
                dictionary[name] = events
            if sheet[key].value == None:
                doneSorting = True
            counter = counter + 1
            key = "D" + str(counter)
        return dictionary
    def sortFreshman(self, sheet):
        dictionary = {}
        counter = 2
        key = "D" + str(counter)
        doneSorting = False
        while doneSorting == False:
            if sheet[key].value == "F":
                nameKey = "A" + str(counter)
                eventKey = "C" + str(counter)
                name = sheet[nameKey].value
                events = sheet[eventKey].value
                dictionary[name] = events
            if sheet[key].value == None:
                doneSorting = True
            counter = counter + 1
            key = "D" + str(counter)
        return dictionary
    def turnToString(self, listOfPeople):
        string = ""
        for i in listOfPeople:
            string = string + i
            if i == listOfPeople[-1]:
                return string
            string = string + ", "
    def updateGoogleSheet(self, sheet, teams):
        teamOne = teams[0]
        teamTwo = teams[1]
        teamThree = teams[2]
        counter = 2
        keyOne = "G" + str(counter)
        for i in teamOne:
            for value in i.values():
                sheet[keyOne].value = self.turnToString(value)
            counter+=1
            keyOne = "G" + str(counter)
        counter = 2
        keyTwo = "H" + str(counter)
        for i in teamTwo:
            for value in i.values():
                sheet[keyTwo].value = self.turnToString(value)
            counter+=1
            keyTwo = "H" + str(counter)
        counter = 2
        keyThree = "I" + str(counter)
        for i in teamThree:
            for value in i.values():
                sheet[keyThree].value = self.turnToString(value)
            counter+=1
            keyThree = "H" + str(counter)
    def dictCleanup(self, team):
        for i in team:
            index = team.index(i)
            for key,value in i.items():
                newKey = ''.join([i for i in key if not i.isdigit()])
                newDict = {newKey:value}
                team[index] = newDict
        return team
    def putLeftOverInEvent(self, team, listOfEvents, name):
        putInEvent = False
        for i in listOfEvents:
            #loop through the list of dictionaries
            if putInEvent == True:
                break
            for p in team:
                index = team.index(p)
                if(putInEvent == True):
                    break
                #get the dictionary
                for key,value in p.items():
                    #check to see if the key of the dictionary (event name) matches the desired event
                    if key == i:
                        #get the max team size
                        eventCounter = 2
                        eventKey = "F" + str(eventCounter)
                        while sheet[eventKey].value != i:
                            eventCounter+=1
                            eventKey = "F" + str(eventCounter)
                        numKey = "J" + str(eventCounter)
                        maxTeamSize = sheet[numKey].value
                        currentParticipants = value
                        if maxTeamSize != len(currentParticipants):
                            #there is an opening on the team so the leftover person should be added to the team
                            currentParticipants.append(name)
                            newTeam = {key:currentParticipants}
                            team[index] = newTeam
                            putInEvent = True
                            break
        #a slot in their desired events was not available so they will be put in a random event
        if putInEvent == False:
            for z in team:
                index = team.index(z)
                if putInEvent == True:
                    break
                for key,value in z.items():
                    eventCounter = 2
                    eventKey = "F" + str(eventCounter)
                    while sheet[eventKey].value != key:
                        eventCounter+=1
                        eventKey = "F" + str(eventCounter)
                    numKey = "J" + str(eventCounter)
                    maxTeamSize = sheet[numKey].value
                    currentTeam = value
                    if maxTeamSize != len(value):
                        currentTeam.append(name)
                        newTeam = {key:currentTeam}
                        team[index] = newTeam
                        putInEvent = True
                    if putInEvent == True:
                        break
    def handleLeftOvers(self, teams, leftoverSeniors, leftoverJuniors, leftoverSophomores, leftoverFreshman, teamAssignment, sheet):
        teamOne = teams[0]
        teamTwo = teams[1]
        teamThree = teams[2]
        for i in leftoverSeniors:
            counter = 2
            nameKey = "A" + str(counter)
            while sheet[nameKey].value != i:
                counter+=1
                nameKey = "A" + str(counter)
            desiredEventsKey = "C" + str(counter)
            desiredEvents = sheet[desiredEventsKey].value
            #get what events the leftover person wants to be in
            listOfEvents = desiredEvents.split(", ")
            if teamAssignment[i] == 1:
                team = teamOne
            if teamAssignment[i] == 2:
                team = teamTwo
            if teamAssignment[i] == 3:
                team = teamThree
            self.putLeftOverInEvent(team, listOfEvents, i)
        for i in leftoverJuniors:
            counter = 2
            nameKey = "A" + str(counter)
            while sheet[nameKey].value != i:
                counter+=1
                nameKey = "A" + str(counter)
            desiredEventsKey = "C" + str(counter)
            desiredEvents = sheet[desiredEventsKey].value
            #get what events the leftover person wants to be in
            listOfEvents = desiredEvents.split(", ")
            if teamAssignment[i] == 1:
                team = teamOne
            if teamAssignment[i] == 2:
                team = teamTwo
            if teamAssignment[i] == 3:
                team = teamThree
            self.putLeftOverInEvent(team, listOfEvents, i)
        for i in leftoverSophomores:
            counter = 2
            nameKey = "A" + str(counter)
            while sheet[nameKey].value != i:
                counter+=1
                nameKey = "A" + str(counter)
            desiredEventsKey = "C" + str(counter)
            desiredEvents = sheet[desiredEventsKey].value
            #get what events the leftover person wants to be in
            listOfEvents = desiredEvents.split(", ")
            if teamAssignment[i] == 1:
                team = teamOne
            if teamAssignment[i] == 2:
                team = teamTwo
            if teamAssignment[i] == 3:
                team = teamThree
            self.putLeftOverInEvent(team, listOfEvents, i)
        for i in leftoverFreshman:
            counter = 2
            nameKey = "A" + str(counter)
            while sheet[nameKey].value != i:
                counter+=1
                nameKey = "A" + str(counter)
            desiredEventsKey = "C" + str(counter)
            desiredEvents = sheet[desiredEventsKey].value
            #get what events the leftover person wants to be in
            listOfEvents = desiredEvents.split(", ")
            if teamAssignment[i] == 1:
                team = teamOne
            if teamAssignment[i] == 2:
                team = teamTwo
            if teamAssignment[i] == 3:
                team = teamThree
            self.putLeftOverInEvent(team, listOfEvents, i)
        return [teamOne, teamTwo, teamThree]
    def separateTeams(self, teams, leftovers, sheet):
        teamOne = []
        teamTwo = []
        teamThree = []
        for team in teams:
            if '2' in team:
                dictionaryTwo = {team:teams[team]}
                teamTwo.append(dictionaryTwo)
            elif '3' in team:
                dictionaryThree = {team:teams[team]}
                teamThree.append(dictionaryThree)
            else:
                dictionary = {team:teams[team]}
                teamOne.append(dictionary)
            names = self.getAllParticipants(sheet)
            nameCountOne = {}
            nameCountTwo = {}
            nameCountThree = {}
            for i in names:
                counter = 0
                for t in teamOne:
                    for g in t:
                        for f in t[g]:
                            if i == f:
                                counter+=1
                                nameCountOne[i] = counter
                if counter == 0:
                    nameCountOne[i] = 0
                counter = 0
                for t in teamTwo:
                    for g in t:
                        for f in t[g]:
                            if i == f:
                                counter+=1
                                nameCountTwo[i] = counter
                if counter == 0:
                    nameCountTwo[i] = 0
                counter = 0
                for t in teamThree:
                    for g in t:
                        for f in t[g]:
                            if i == f:
                                counter+=1
                                nameCountThree[i] = counter
                if counter == 0:
                    nameCountThree[i] = 0
                counter = 0
            teamAssignment = {}
            for i in names:
                x = nameCountOne[i]
                y = nameCountTwo[i]
                z = nameCountThree[i]
                if max(x,y,z) == x:
                    teamAssignment[i] = 1
                if max(x,y,z) == y:
                    teamAssignment[i] = 2
                if max(x,y,z) == z:
                    teamAssignment[i] = 3
            for i in teamOne:
                newTeam = []
                index = teamOne.index(i)
                for key in i.keys():
                    eventName = key
                for values in i.values():
                    eventParticipants = values
                for p in eventParticipants:
                    if teamAssignment[p] == 1:
                        newTeam.append(p)
                    else:
                        leftovers.append(p)
                newDict = {}
                newDict[eventName] = newTeam
                teamOne[index] = newDict
            for i in teamTwo:
                newTeam = []
                index = teamTwo.index(i)
                for key in i.keys():
                    eventName = key
                for values in i.values():
                    eventParticipants = values
                for p in eventParticipants:
                    if teamAssignment[p] == 2:
                        newTeam.append(p)
                    else:
                        leftovers.append(p)
                newDict = {}
                newDict[eventName] = newTeam
                teamTwo[index] = newDict
            for i in teamThree:
                newTeam = []
                index = teamThree.index(i)
                for key in i.keys():
                    eventName = key
                for values in i.values():
                    eventParticipants = values
                for p in eventParticipants:
                    if teamAssignment[p] == 3:
                        newTeam.append(p)
                    else:
                        leftovers.append(p)
                newDict = {}
                newDict[eventName] = newTeam
                teamThree[index] = newDict
        return [teamOne, teamTwo, teamThree], leftovers, teamAssignment
    def getAllParticipants(self, sheet):
        counter = 2
        key = "A" + str(counter)
        names = []
        while sheet[key].value != None:
            names.append(sheet[key].value)
            counter+=1
            key = "A" + str(counter)
        return names
    def sortByYear(self, sheet, leftovers):
        leftoverSeniors = []
        leftoverJuniors = []
        leftoverSophomores = []
        leftoverFreshman = []
        for i in leftovers:
            counter = 2
            key = "A" + str(counter)
            while sheet[key].value != None:
                if sheet[key].value == i:
                    gradeKey = "D" + str(counter)
                    grade = sheet[gradeKey].value
                    if grade == "Sr":
                        leftoverSeniors.append(i)
                    if grade == "J":
                        leftoverJuniors.append(i)
                    if grade == "S":
                        leftoverSophomores.append(i)
                    if grade == "F":
                        leftoverFreshman.append(i)
                counter+=1
                key = "A" + str(counter)
        return leftoverSeniors, leftoverJuniors, leftoverSophomores, leftoverFreshman
def getFileName():
    print("enter in the name of the .xlsx file that you want to generate groups and teams for")
    userInput = input(".xlsx file with team data: ")
    return userInput
def setupEventList(sheet):
    event1 = {"Event name": sheet["F2"].value}
    event2 = {"Event name": sheet["F3"].value}
    event3 = {"Event name": sheet["F4"].value}
    event4 = {"Event name": sheet["F5"].value}
    event5 = {"Event name": sheet["F6"].value}
    event6 = {"Event name": sheet["F7"].value}
    event7 = {"Event name": sheet["F8"].value}
    event8 = {"Event name": sheet["F9"].value}
    event9 = {"Event name": sheet["F10"].value}
    event10 = {"Event name": sheet["F11"].value}
    event11 = {"Event name": sheet["F12"].value}
    event12 = {"Event name": sheet["F13"].value}
    event13 = {"Event name": sheet["F14"].value}
    event14 = {"Event name": sheet["F15"].value}
    event15 = {"Event name": sheet["F16"].value}
    event16 = {"Event name": sheet["F17"].value}
    event17 = {"Event name": sheet["F18"].value}
    event18 = {"Event name": sheet["F19"].value}
    event19 = {"Event name": sheet["F20"].value}
    event20 = {"Event name": sheet["F21"].value}
    event21 = {"Event name": sheet["F22"].value}
    event22 = {"Event name": sheet["F23"].value}
    event23 = {"Event name": sheet["F24"].value}
    return event1, event2, event3, event4, event5, event6, event7, event8, event9, event10, event11, event12, event13, event14, event15, event16, event17, event18, event19, event20, event21, event22, event23

filename = getFileName()
algorithm = Grouping_Algorithm()
sheet, workbook, xlfile = algorithm.getWorkbook(filename)
if sheet:
    seniors = algorithm.sortSeniors(sheet=sheet)
    juniors = algorithm.sortJuniors(sheet=sheet)
    sophomores = algorithm.sortSophomores(sheet=sheet)
    freshmen = algorithm.sortFreshman(sheet=sheet)
    event1, event2, event3, event4, event5, event6, event7, event8, event9, event10, event11, event12, event13, event14, event15, event16, event17, event18, event19, event20, event21, event22, event23 = setupEventList(sheet)
    eventList = [event1, event2, event3, event4, event5, event6, event7, event8, event9, event10, event11, event12, event13, event14, event15, event16, event17, event18, event19, event20, event21, event22, event23]
    algorithm.putInEvent(seniors, eventList)
    algorithm.putInEvent(juniors, eventList)
    algorithm.putInEvent(sophomores, eventList)
    algorithm.putInEvent(freshmen, eventList)
    numTeams = algorithm.getNumOfTeams(sheet)
    teams, leftovers = algorithm.putInTeams(sheet, eventList, numTeams)
    teams, leftovers, teamAssignment = algorithm.separateTeams(teams, leftovers, sheet)
    teams[1] = algorithm.dictCleanup(teams[1])
    teams[2] = algorithm.dictCleanup(teams[2])
    leftoverSeniors, leftoverJuniors, leftoverSophomores, leftoverFreshman = algorithm.sortByYear(sheet, leftovers)
    teams = algorithm.handleLeftOvers(teams, leftoverSeniors, leftoverJuniors, leftoverSophomores, leftoverFreshman, teamAssignment, sheet)
    algorithm.updateGoogleSheet(sheet, teams)
    workbook.save(xlfile)